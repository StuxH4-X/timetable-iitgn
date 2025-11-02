from flask import Flask, render_template, request, send_file
import pandas as pd
import json
import io
import re

from openpyxl import load_workbook
from openpyxl.styles import Alignment

app = Flask(__name__)

FILE_PATH = "Timetable, Sem-II, 2025-26.xlsx"

@app.route("/", methods=["GET"])
def main_page():
    try:
        
        df = pd.read_excel(FILE_PATH, sheet_name="Time table", na_filter=False)

        selected_columns = [
            "Course Number", "Course Name", "L", "T", "P", "C",
            "Name of the Instructors and Tutors", "Lecture", "Tutorial", "Lab", "HSS/BS elective"
        ]
        df = df[selected_columns]

        
        df = df[df['Name of the Instructors and Tutors'] != '']
        df['priority'] = -1
        df['searched'] = 0

        
        table_data = df.reset_index().to_dict(orient="records")

        return render_template("index.html", table_data=table_data)
    except Exception as e:
        return f"Error loading file: {e}", 500
    
selected_courses = []  

@app.route("/save_selected_courses", methods=["POST"])
def save_selected_courses():
    global selected_courses
    try:
        selected_courses = json.loads(request.data)
        return "Success", 200
    except Exception as e:
        return f"Error saving courses: {e}", 500

@app.route("/timetable", methods=["GET"])
def timetable_page():
    global selected_courses
    try:
        df = pd.read_excel(FILE_PATH, sheet_name="Time Slots", na_filter=False)  #, skiprows=[0, 5]

        time_slots_data = df.to_dict(orient="records")

        return render_template(
            "index2.html", 
            selected_data=selected_courses, 
            time_slots_data=time_slots_data
        )
    except Exception as e:
        return f"Error loading timetable: {e}", 500


@app.route('/process-excel', methods=['POST'])
def process_excel():
    try:
        
        uploaded_file = request.files['file']
        if not uploaded_file:
            return "No file uploaded", 400

        workbook = load_workbook(uploaded_file)
          
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):  
                        cell.value = re.sub(r'\)(.)', r')\1\n', cell.value)
                    if cell.value: 
                        cell.alignment = Alignment(wrap_text=True)

        
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="processed_file.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        return f"Error processing Excel file: {str(e)}", 500

